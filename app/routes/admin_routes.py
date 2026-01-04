"""Admin routes for managing items and orders.

Includes endpoints for:
- Creating, deleting, and toggling items
- Viewing and deleting orders
- Exporting orders in various formats
"""

import io
import logging
import uuid
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO

import openpyxl
from app.settings.config import templates
from fastapi import APIRouter, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from openpyxl import Workbook

router = APIRouter(prefix='/admin', tags=['admin'])


@router.post('/items/create')
async def create_item(request: Request, name: str = Form(...), price: float = Form(...), ttl: int = Form(...)):
    """Create a new item. Requires admin privileges."""
    cashier_id = request.session.get('cashier_id')
    async with request.app.state.db.acquire() as conn:
        cashier = await conn.fetchrow('SELECT is_admin FROM cashiers WHERE id = $1', cashier_id)
        if not cashier or not cashier['is_admin']:
            return RedirectResponse('/', status_code=302)

        await conn.execute(
            'INSERT INTO items (name, price, ttl, active) VALUES ($1, $2, $3, TRUE)',
            name, price, ttl
        )

    return RedirectResponse('/admin/items', status_code=302)


@router.post('/items/delete')
async def delete_item(request: Request, item_id: str = Form(...)):
    """Delete an item by ID. Requires admin privileges."""
    cashier_id = request.session.get('cashier_id')
    async with request.app.state.db.acquire() as conn:
        cashier = await conn.fetchrow('SELECT is_admin FROM cashiers WHERE id = $1', cashier_id)
        if not cashier or not cashier['is_admin']:
            return RedirectResponse('/', status_code=302)

        await conn.execute('DELETE FROM items WHERE id = $1', uuid.UUID(item_id))

    return RedirectResponse('/admin/items', status_code=302)


@router.get('/items', response_class=HTMLResponse)
async def admin_items(request: Request):
    """View all items. Requires admin privileges."""
    cashier_id = request.session.get('cashier_id')
    if not cashier_id:
        return RedirectResponse('/', status_code=302)

    async with request.app.state.db.acquire() as conn:
        cashier = await conn.fetchrow('SELECT is_admin FROM cashiers WHERE id = $1', cashier_id)
        if not cashier or not cashier['is_admin']:
            return RedirectResponse('/', status_code=302)

        items = await conn.fetch('SELECT id, name, active FROM items ORDER BY name')

    return templates.TemplateResponse('admin_items.html', {'request': request, 'items': items})


@router.post('/items/toggle')
async def toggle_item_activity(request: Request):
    """Toggle multiple item active statuses. Requires admin privileges."""
    cashier_id = request.session.get('cashier_id')
    db = request.app.state.db
    
    async with db.acquire() as conn:
        cashier = await conn.fetchrow('SELECT is_admin FROM cashiers WHERE id = $1', cashier_id)
        if not cashier or not cashier['is_admin']:
            return RedirectResponse('/', status_code=302)

        # Get all form data
        form_data = await request.form()
        
        # Get all items first to know which ones should be active
        items = await conn.fetch('SELECT id FROM items ORDER BY id')
        
        for item in items:
            item_id = item['id']
            # Check if checkbox was submitted (True) or not (False)
            is_active = f'active_{item_id}' in form_data
            
            await conn.execute(
                'UPDATE items SET active = $1 WHERE id = $2',
                is_active,
                item_id
            )

    return RedirectResponse('/admin/items', status_code=302)



@router.get('/orders', response_class=HTMLResponse)
async def admin_orders(request: Request):
    """View orders with optional filtering by date and address. Requires admin privileges."""
    cashier_id = request.session.get('cashier_id')
    if not cashier_id:
        return RedirectResponse('/', status_code=302)

    db = request.app.state.db

    # Check admin status
    async with db.acquire() as conn:
        cashier = await conn.fetchrow('SELECT is_admin FROM cashiers WHERE id = $1', cashier_id)
        if not cashier or not cashier['is_admin']:
            return RedirectResponse('/', status_code=302)

    order_for_date_str = request.query_params.get('order_for_date')
    address_filter = request.query_params.get('address')

    where_clauses: list = []
    query_args: list = []
    param_counter = 1

    if order_for_date_str:
        try:
            order_for_date = datetime.strptime(order_for_date_str, '%Y-%m-%d').date()
            where_clauses.append(f'o.order_for = ${param_counter}')
            query_args.append(order_for_date)
            param_counter += 1
        except ValueError:
            logging.warning(f'Invalid date format received: {order_for_date_str}')

    if address_filter:
        where_clauses.append(f'o.address ILIKE ${param_counter}')
        query_args.append(f'%{address_filter}%')
        param_counter += 1

    where_sql = ''
    if where_clauses:
        where_sql = ' WHERE ' + ' AND '.join(where_clauses)

    async with db.acquire() as conn:
        rows = await conn.fetch(f"""
            SELECT
                o.id AS order_id,
                o.order_for,
                o.created,
                o.address,
                c.full_name AS cashier_name,
                oi.quantity,
                i.name AS item_name
            FROM orders o
            JOIN cashiers c ON o.cashier_id = c.id
            JOIN orders_items oi ON oi.order_id = o.id
            JOIN items i ON oi.item_id = i.id
            {where_sql}
            ORDER BY o.order_for ASC, o.created ASC
        """, *query_args)

    grouped_orders = {}
    for row in rows:
        date_key = row['order_for']
        o_id = str(row['order_id'])
        if date_key not in grouped_orders:
            grouped_orders[date_key] = {}
        if o_id not in grouped_orders[date_key]:
            grouped_orders[date_key][o_id] = {
                'id': o_id,
                'created': row['created'],
                'address': row['address'],
                'cashier_name': row['cashier_name'],
                'items': [],
            }
        grouped_orders[date_key][o_id]['items'].append({
            'name': row['item_name'],
            'quantity': row['quantity'],
        })

    return templates.TemplateResponse('admin_orders.html', {
        'request': request,
        'grouped_orders': grouped_orders,
        'order_for': order_for_date_str,
        'address': address_filter
    })


@router.get('/orders/export')
async def export_orders(address: str = None, request: Request = None):
    """Export orders to Excel, optionally filtered by address. Requires admin privileges."""
    db = request.app.state.db
    async with db.acquire() as conn:
        query = """
            SELECT o.id, o.created, o.address, c.full_name AS cashier_name
            FROM orders o
            JOIN cashiers c ON o.cashier_id = c.id
        """
        params = []
        if address:
            query += ' WHERE o.address ILIKE $1'
            params.append(f'%{address}%')
        query += ' ORDER BY o.created DESC'

        orders = await conn.fetch(query, *params)
        logging.info(f'Orders fetched for export: {len(orders)}')
        if not orders:
            return {'detail': 'No orders found'}

        wb = Workbook()
        wb.remove(wb.active)

        # Cache worksheets by address
        sheet_map = {}

        for order in orders:
            addr = order['address'] or 'Unknown'
            sheet_name = addr[:31]  # Excel sheet name limit

            # Create or reuse sheet
            if sheet_name not in sheet_map:
                ws = wb.create_sheet(title=sheet_name)
                ws.append(['Order ID', 'Created', 'Address', 'Cashier', 'Item Name', 'Quantity', 'Price'])
                sheet_map[sheet_name] = ws
            else:
                ws = sheet_map[sheet_name]

            items = await conn.fetch("""
                SELECT oi.quantity, i.name, i.price
                FROM orders_items oi
                JOIN items i ON i.id = oi.item_id
                WHERE oi.order_id = $1
            """, order['id'])

            for item in items:
                ws.append([
                    str(order['id']),
                    order['created'].strftime('%Y-%m-%d %H:%M:%S'),
                    order['address'],
                    order['cashier_name'],
                    item['name'],
                    item['quantity'],
                    float(item['price'])
                ])

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        headers = {
            'Content-Disposition': 'attachment; filename=orders.xlsx'
        }

        return StreamingResponse(stream, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@router.get('/export/by_address')
async def export_by_address(order_for: date, request: Request):
    """Export orders by address for a specific date. Requires admin privileges."""
    db = request.app.state.db
    async with db.acquire() as conn:
        query = """
            SELECT
                o.address,
                i.name AS item_name,
                oi.quantity
            FROM orders o
            JOIN orders_items oi ON o.id = oi.order_id
            JOIN items i ON oi.item_id = i.id
            WHERE o.order_for = $1
            ORDER BY o.address
        """
        rows = await conn.fetch(query, order_for)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Group by address
    grouped = defaultdict(list)
    for row in rows:
        grouped[row['address']].append((row['item_name'], row['quantity']))

    for address, items in grouped.items():
        ws = wb.create_sheet(title=address[:31])  # Excel max 31 chars
        ws.append(['Товар', 'Количество'])
        for name, qty in items:
            ws.append([name, qty])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={
        'Content-Disposition': f'attachment; filename=by_address_{order_for}.xlsx'
    })


@router.get('/export/all_items')
async def export_all_items(order_for: date, request: Request):
    """Export all items ordered for a specific date. Requires admin privileges."""
    db = request.app.state.db
    async with db.acquire() as conn:
        query = """
            SELECT
                o.id AS order_id,
                i.name AS item_name,
                oi.quantity,
                o.address,
                o.order_for
            FROM orders o
            JOIN orders_items oi ON o.id = oi.order_id
            JOIN items i ON oi.item_id = i.id
            WHERE o.order_for = $1
            ORDER BY o.id, i.name;
        """
        rows = await conn.fetch(query, order_for)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'All Items by Order'

    # Set headers
    ws.append([
        'Order ID',
        'Name of Item',
        'Quantity',
        'Address',
        'Order Date'
    ])

    # Append data rows
    for row in rows:
        ws.append([
            str(row['order_id']),
            row['item_name'],
            row['quantity'],
            row['address'],
            row['order_for']
        ])

    # Prepare response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename=all_items_{order_for}.xlsx'
        }
    )


@router.post('/orders/delete')
async def delete_order(request: Request, order_id: str = Form(...)):
    """Delete an order and its associated items. Requires admin privileges."""
    cashier_id = request.session.get('cashier_id')
    db = request.app.state.db

    logging.info(f"Attempting to delete order with ID: '{order_id}' by cashier ID: {cashier_id}")

    # Defensive check: empty or missing order_id
    if not order_id or not order_id.strip():
        logging.error('No order_id provided in form submission.')
        return HTMLResponse('Missing order ID.', status_code=400)

    try:
        order_uuid = uuid.UUID(order_id.strip())
        logging.info(f'Successfully parsed order_id to UUID: {order_uuid}')
    except ValueError:
        logging.error(f'Invalid UUID format for order_id: {order_id}')
        return HTMLResponse(f'Invalid order ID format: {order_id}', status_code=400)

    async with db.acquire() as conn:
        # Check if the user is an admin
        cashier = await conn.fetchrow('SELECT is_admin FROM cashiers WHERE id = $1', cashier_id)
        if not cashier or not cashier['is_admin']:
            logging.warning(f'Unauthorized delete attempt by cashier ID: {cashier_id}')
            return RedirectResponse('/', status_code=302)

        try:
            # Start transaction
            async with conn.transaction():
                # Delete items
                items_deleted = await conn.execute('DELETE FROM orders_items WHERE order_id = $1', order_uuid)
                logging.info(f'Deleted from orders_items: {items_deleted}')

                # Delete order
                order_deleted = await conn.execute('DELETE FROM orders WHERE id = $1', order_uuid)
                logging.info(f'Deleted from orders: {order_deleted}')

            if order_deleted and 'DELETE' in order_deleted:
                logging.info(f'Successfully deleted order {order_uuid}')
                return RedirectResponse('/admin/orders', status_code=302)
            else:
                logging.warning(f'No order deleted with ID {order_uuid}')
                return HTMLResponse('Order not found or already deleted.', status_code=404)

        except Exception as e:
            logging.error(f'Error deleting order {order_id}: {e}', exc_info=True)
            return HTMLResponse('Internal server error while deleting the order.', status_code=500)
