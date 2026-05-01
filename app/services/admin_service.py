#DEPECATED
import io
import uuid
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO

import openpyxl
from openpyxl import Workbook

from app.infrastructure.uow import AsyncpgUnitOfWork


class AdminService:
    @staticmethod
    def _build_live_orders_payload(
        rows: list[dict], all_item_names: list[str], *, max_days: int | None = None
    ) -> dict:
        today = date.today()
        days_map: dict[date, dict] = {}

        for r in rows:
            day = r['order_for']
            day_block = days_map.setdefault(
                day,
                {
                    'date': day.isoformat(),
                    'is_today': day == today,
                    'total_orders': 0,
                    'totals_map': defaultdict(int, {name: 0 for name in all_item_names}),
                    'orders_map': {},
                    'shops_map': {},
                },
            )

            oid = r['order_id']
            order_block = day_block['orders_map'].setdefault(
                oid,
                {
                    'id': oid,
                    'created': r['created'],
                    'created_hhmm': r['created'].strftime('%H:%M:%S'),
                    'address': r['address'],
                    'cashier_name': r['cashier_name'],
                    'shop_id': r['shop_id'],
                    'items': [],
                },
            )
            order_block['items'].append(
                {'name': r['item_name'], 'quantity': int(r['quantity'])}
            )

            shop_key = r['address'] or (r['shop_id'] or 'Unknown shop')
            shop_block = day_block['shops_map'].setdefault(
                shop_key, {'shop': shop_key, 'orders': {}}
            )
            shop_block['orders'][oid] = order_block

            day_block['totals_map'][r['item_name']] += int(r['quantity'])

        sorted_days = sorted(days_map.keys(), reverse=True)
        if max_days is not None:
            sorted_days = sorted_days[:max_days]

        days: list[dict] = []
        for day in sorted_days:
            block = days_map[day]
            orders_sorted = sorted(
                block['orders_map'].values(), key=lambda x: x['created'], reverse=True
            )
            block['total_orders'] = len(orders_sorted)

            totals = [
                {'name': name, 'quantity': qty}
                for name, qty in sorted(block['totals_map'].items())
            ]

            shops = []
            for shop_name, shop_data in sorted(block['shops_map'].items()):
                shop_orders = sorted(
                    shop_data['orders'].values(),
                    key=lambda x: x['created'],
                    reverse=True,
                )
                shops.append({'shop': shop_name, 'orders': shop_orders})

            for o in orders_sorted:
                o['created'] = o['created'].isoformat()

            days.append(
                {
                    'date': block['date'],
                    'is_today': block['is_today'],
                    'total_orders': block['total_orders'],
                    'totals': totals,
                    'shops': shops,
                }
            )

        return {'days': days, 'generated_at': datetime.utcnow().isoformat()}

    @staticmethod
    async def create_item(
        *,
        uow: AsyncpgUnitOfWork,
        name: str,
        category_id: uuid.UUID | None,
    ) -> None:
        assert uow.items is not None
        await uow.items.create(
            name=name, category_id=category_id
        )

    @staticmethod
    async def delete_item(*, uow: AsyncpgUnitOfWork, item_id: uuid.UUID) -> None:
        assert uow.items is not None
        await uow.items.delete(item_id=item_id)

    @staticmethod
    async def list_items(*, uow: AsyncpgUnitOfWork) -> list[dict]:
        assert uow.items is not None
        return await uow.items.list_for_admin()

    @staticmethod
    async def toggle_items(
        *,
        uow: AsyncpgUnitOfWork,
        active_map: dict[uuid.UUID, bool],
        category_map: dict[uuid.UUID, uuid.UUID | None],
    ) -> None:
        assert uow.items is not None
        for item_id, is_active in active_map.items():
            await uow.items.update_admin_fields(
                item_id=item_id,
                active=is_active,
                category_id=category_map.get(item_id),
            )

    @staticmethod
    async def list_categories(*, uow: AsyncpgUnitOfWork) -> list[dict]:
        assert uow.items is not None
        return await uow.items.list_categories()

    @staticmethod
    async def create_category(*, uow: AsyncpgUnitOfWork, name: str) -> None:
        assert uow.items is not None
        await uow.items.create_category(name=name)

    @staticmethod
    async def rename_category(
        *, uow: AsyncpgUnitOfWork, category_id: uuid.UUID, name: str
    ) -> None:
        assert uow.items is not None
        await uow.items.rename_category(category_id=category_id, name=name)

    @staticmethod
    async def get_orders(
        *, uow: AsyncpgUnitOfWork, order_for: str | None, address: str | None
    ) -> dict:
        assert uow.orders is not None

        order_for_date: date | None = None
        if order_for:
            try:
                order_for_date = datetime.strptime(order_for, '%Y-%m-%d').date()
            except ValueError:
                order_for_date = None

        rows = await uow.orders.admin_rows(order_for=order_for_date, address=address)

        grouped: dict = {}
        for r in rows:
            d = r['order_for']
            oid = r['order_id']
            grouped.setdefault(d, {}).setdefault(
                oid,
                {
                    'id': oid,
                    'created': r['created'],
                    'address': r['address'],
                    'cashier_name': r['cashier_name'],
                    'items': [],
                },
            )['items'].append({'name': r['item_name'], 'quantity': r['quantity']})

        return grouped

    @staticmethod
    async def delete_order(*, uow: AsyncpgUnitOfWork, order_id: uuid.UUID) -> bool:
        assert uow.orders is not None
        return await uow.orders.delete_order(order_id=order_id)

    @staticmethod
    async def export_orders(
        *, uow: AsyncpgUnitOfWork, address: str | None
    ) -> io.BytesIO:
        assert uow.orders is not None

        orders = await uow.orders.export_orders_list(address=address)

        wb = Workbook()
        wb.remove(wb.active)

        sheets = {}
        for o in orders:
            addr = o['address'] or 'Unknown'
            name = addr[:31]
            ws = sheets.setdefault(name, wb.create_sheet(title=name))
            if ws.max_row == 1:
                ws.append(
                    [
                        'Order ID',
                        'Created',
                        'Address',
                        'Cashier',
                        'Item',
                        'Qty',
                    ]
                )

            items = await uow.orders.export_order_items(order_id=uuid.UUID(o['id']))
            for it in items:
                ws.append(
                    [
                        o['id'],
                        o['created'].strftime('%Y-%m-%d %H:%M:%S'),
                        o['address'],
                        o['cashier_name'],
                        it['name'],
                        it['quantity'],
                    ]
                )

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    @staticmethod
    async def export_by_address(*, uow: AsyncpgUnitOfWork, order_for: date) -> BytesIO:
        assert uow.orders is not None

        rows = await uow.orders.export_by_address_rows(order_for=order_for)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        grouped = defaultdict(list)
        for r in rows:
            grouped[r['address']].append((r['name'], r['quantity']))

        for addr, items in grouped.items():
            ws = wb.create_sheet(title=(addr or 'Unknown')[:31])
            ws.append(['Item', 'Quantity'])
            for n, q in items:
                ws.append([n, q])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    async def export_all_items(*, uow: AsyncpgUnitOfWork, order_for: date) -> BytesIO:
        assert uow.orders is not None

        rows = await uow.orders.export_by_address_rows(order_for=order_for)

        totals: dict[str, int] = defaultdict(int)
        for r in rows:
            name = r["name"]
            qty = int(r["quantity"])
            totals[name] += qty

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All items"
        ws.append(["Item", "Quantity"])

        for name in sorted(totals.keys()):
            ws.append([name, totals[name]])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    async def get_live_orders_payload(*, uow: AsyncpgUnitOfWork) -> dict:
        assert uow.orders is not None
        assert uow.items is not None
        rows = await uow.orders.admin_rows_live()
        all_items = await uow.items.list_for_admin()
        all_item_names = sorted({str(i['name']) for i in all_items})
        return AdminService._build_live_orders_payload(
            rows, all_item_names, max_days=5
        )

    @staticmethod
    async def get_live_orders_archive_payload(*, uow: AsyncpgUnitOfWork) -> dict:
        assert uow.orders is not None
        assert uow.items is not None
        rows = await uow.orders.admin_rows_live()
        all_items = await uow.items.list_for_admin()
        all_item_names = sorted({str(i['name']) for i in all_items})
        full_payload = AdminService._build_live_orders_payload(rows, all_item_names)
        archive_days = full_payload['days'][5:]
        return {
            'days': archive_days,
            'generated_at': full_payload['generated_at'],
        }

    @staticmethod
    async def export_live_totals(*, uow: AsyncpgUnitOfWork, order_for: date) -> BytesIO:
        assert uow.orders is not None
        assert uow.items is not None

        all_items = await uow.items.list_for_admin()
        all_item_names = sorted({str(i['name']) for i in all_items})
        rows = await uow.orders.admin_rows(order_for=order_for, address=None)

        totals: dict[str, int] = {name: 0 for name in all_item_names}
        for r in rows:
            name = str(r['item_name'])
            qty = int(r['quantity'])
            totals[name] = totals.get(name, 0) + qty

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Totals {order_for}"
        ws.append(["Item", "Ordered quantity"])

        for name in sorted(totals.keys()):
            ws.append([name, totals[name]])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
