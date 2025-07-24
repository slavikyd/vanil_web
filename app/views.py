import io
import logging
import uuid
from datetime import date, datetime
from io import BytesIO

import openpyxl
from asyncpg import Record
from fastapi import APIRouter, Form, Request
from fastapi.responses import (HTMLResponse, JSONResponse, RedirectResponse,
                               StreamingResponse)
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from pydantic import BaseModel

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")


if not hasattr(router, "carts"):
    router.carts = {}

def get_cashier_id(request: Request):
    """Retrieve cashier ID from session."""
    return request.session.get("cashier_id")

@router.get("/", response_class=HTMLResponse)
async def index(request: Request):
    cashier_id = get_cashier_id(request)
    if not cashier_id:
        return templates.TemplateResponse("index.html", {"request": request, "cashier_id": None})

    async with request.app.state.db.acquire() as conn:
        # Получаем товары
        items = await conn.fetch("SELECT id, name, price FROM items WHERE active = TRUE ORDER BY name ASC")
        
        # Получаем данные кассира, чтобы узнать is_admin
        cashier_record = await conn.fetchrow("SELECT is_admin FROM cashiers WHERE id = $1", cashier_id)

    items_list_for_template = []
    for item in items:
        items_list_for_template.append({
            "id": str(item["id"]),
            "name": item["name"],
            "price": float(item["price"])
        })

    session_id = request.session.get("session_id")
    if not session_id or session_id not in router.carts:
        session_id = str(uuid.uuid4())
        router.carts[session_id] = {}
        request.session["session_id"] = session_id

    serializable_cart = {str(k): v for k, v in router.carts[session_id].items()}

    is_admin = False
    if cashier_record:
        is_admin = cashier_record["is_admin"]

    return templates.TemplateResponse("index.html", {
        "request": request,
        "items": items_list_for_template,
        "cart": serializable_cart,
        "cashier_id": cashier_id,
        "is_admin": is_admin,
    })

@router.post("/admin/items/create")
async def create_item(request: Request, name: str = Form(...), price: float = Form(...), ttl: int = Form(...)):
    cashier_id = get_cashier_id(request)
    async with request.app.state.db.acquire() as conn:
        cashier = await conn.fetchrow("SELECT is_admin FROM cashiers WHERE id = $1", cashier_id)
        if not cashier or not cashier["is_admin"]:
            return RedirectResponse("/", status_code=302)

        await conn.execute(
            "INSERT INTO items (name, price, ttl, active) VALUES ($1, $2, $3, TRUE)",
            name, price, ttl
        )

    return RedirectResponse("/admin/items", status_code=302)


@router.post("/admin/items/delete")
async def delete_item(request: Request, item_id: str = Form(...)):
    cashier_id = get_cashier_id(request)
    async with request.app.state.db.acquire() as conn:
        cashier = await conn.fetchrow("SELECT is_admin FROM cashiers WHERE id = $1", cashier_id)
        if not cashier or not cashier["is_admin"]:
            return RedirectResponse("/", status_code=302)

        await conn.execute("DELETE FROM items WHERE id = $1", uuid.UUID(item_id))

    return RedirectResponse("/admin/items", status_code=302)

@router.post("/login")
async def login(request: Request, cashier_id: str = Form(...)):
    async with request.app.state.db.acquire() as conn:
        cashier = await conn.fetchrow("SELECT id FROM cashiers WHERE id = $1", cashier_id)
        if not cashier:
            return templates.TemplateResponse("index.html", {"request": request, "cashier_id": None, "error": "Invalid cashier ID"})
    
    request.session["cashier_id"] = cashier_id
    session_id = str(uuid.uuid4())
    router.carts[session_id] = {}
    request.session["session_id"] = session_id
    return RedirectResponse("/", status_code=302)

@router.post("/logout")
async def logout(request: Request):
    session_id = request.session.get("session_id")
    if session_id in router.carts:
        del router.carts[session_id]
    request.session.clear()
    return RedirectResponse("/", status_code=302)

class AddToOrderRequest(BaseModel):
    item_id: str
    quantity: int
    tg_id: str = None

@router.post("/add-to-order")
async def add_to_order(request: Request, order_data: AddToOrderRequest):
    cashier_id = get_cashier_id(request)
    if not cashier_id:
        return JSONResponse({"error": "Unauthorized: No cashier logged in"}, status_code=401)

    session_id = request.session.get("session_id")
    if not session_id or session_id not in router.carts:
        session_id = str(uuid.uuid4())
        router.carts[session_id] = {}
        request.session["session_id"] = session_id

    cart = router.carts[session_id]

    item_id = order_data.item_id
    quantity = order_data.quantity

    if quantity > 0:
        cart[item_id] = quantity
    else:
        cart.pop(item_id, None)

    async with request.app.state.db.acquire() as conn:
        items_from_db = await conn.fetch("SELECT id, name, price FROM items ORDER BY name ASC")

    items_list_for_json = []
    for item in items_from_db:
        items_list_for_json.append({
            "id": str(item["id"]),
            "name": item["name"],
            "price": float(item["price"])
        })

    serializable_cart = {str(k): v for k, v in cart.items()}

    return JSONResponse({"cart": serializable_cart, "items_data": items_list_for_json})

@router.post("/place_order")
async def place_order(request: Request, tg_id: str = Form(...), order_for: str = Form(...)):
    cashier_id = get_cashier_id(request)
    if not cashier_id:
        return RedirectResponse("/", status_code=302)

    db = request.app.state.db

    shop = await db.fetchrow("SELECT id FROM shops WHERE id = $1", tg_id)
    if not shop:
        logging.error(f"Shop with tg_id {tg_id} not found when placing order.")
        return HTMLResponse("Shop not registered", status_code=400)

    cashier = await db.fetchrow("SELECT id FROM cashiers WHERE id = $1", cashier_id)
    async with request.app.state.db.acquire() as conn:
        address: list[Record] = await conn.fetch("SELECT * from shops where id = $1", tg_id)

    if not cashier:
        logging.error(f"Cashier with ID {cashier_id} not found when placing order.")
        return HTMLResponse("Cashier not registered", status_code=400)

    order_id = uuid.uuid4()

    # Convert order_for from string to date object
    try:
        order_for_date = datetime.strptime(order_for, "%Y-%m-%d").date()
    except ValueError as e:
        logging.error(f"Invalid date format for order_for: {order_for}")
        return HTMLResponse(f"Invalid date format for order_for: {order_for}", status_code=400)

    try:
        async with db.acquire() as conn:
            await conn.execute(
                """
                INSERT INTO orders (id, cashier_id, shop_id, address, order_for)
                VALUES ($1, $2, $3, $4, $5)
                """,
                order_id, cashier_id, tg_id, address[0]['address'], order_for_date
            )

            session_id = request.session.get("session_id")
            cart = router.carts.get(session_id, {})

            if not cart:
                logging.warning(f"Attempted to place an empty order for cashier {cashier_id}, session {session_id}")
                return HTMLResponse("Your cart is empty. Nothing to order.", status_code=400)

            for item_id, qty in cart.items():
                if qty > 0:
                    await conn.execute(
                        "INSERT INTO orders_items (order_id, item_id, quantity) VALUES ($1, $2, $3)",
                        order_id, uuid.UUID(item_id), qty
                    )

        if session_id in router.carts:
            del router.carts[session_id]

        return RedirectResponse("/", status_code=302)
    except Exception as e:
        logging.error(f"Error placing order: {e}", exc_info=True)
        return HTMLResponse(f"An error occurred while placing your order: {e}", status_code=500)


@router.get("/orders", response_class=HTMLResponse)
async def orders_view(request: Request):
    cashier_id = get_cashier_id(request)
    if not cashier_id:
        return RedirectResponse("/", status_code=302)

    async with request.app.state.db.acquire() as conn:
        orders = await conn.fetch("""
            SELECT o.id, o.created, o.address, c.id as cashier_id, s.id as shop_id, c.is_admin as is_admin, c.full_name AS cashier_name
            FROM orders o
            JOIN cashiers c ON o.cashier_id = c.id
            JOIN shops s on o.shop_id = s.id
            ORDER BY o.created DESC;
        """)

        orders_data = []
        for order in orders:
            items = await conn.fetch("""
                SELECT oi.quantity, i.name, i.price
                FROM orders_items oi
                JOIN items i ON i.id = oi.item_id
                WHERE oi.order_id = $1
            """, order["id"])

            serializable_items = []
            for item in items:
                serializable_items.append({
                    "quantity": item["quantity"],
                    "name": item["name"],
                    "price": float(item["price"]),
                })

            orders_data.append({
                "id": str(order["id"]),
                "created": order["created"],
                "address": order["address"],
                "cashier_name": order["cashier_name"],
                "items": serializable_items,
                "cashier_id": str(order['cashier_id']),
                "shop_id": str(order['shop_id']),
            })

    return templates.TemplateResponse("orders.html", {"request": request, "orders": orders_data})




@router.get("/admin/items", response_class=HTMLResponse)
async def admin_items(request: Request):
    cashier_id = get_cashier_id(request)
    if not cashier_id:
        return RedirectResponse("/", status_code=302)

    async with request.app.state.db.acquire() as conn:
        cashier = await conn.fetchrow("SELECT is_admin FROM cashiers WHERE id = $1", cashier_id)
        if not cashier or not cashier["is_admin"]:
            return RedirectResponse("/", status_code=302)

        items = await conn.fetch("SELECT id, name, active FROM items ORDER BY name")

    return templates.TemplateResponse("admin_items.html", {"request": request, "items": items})


@router.post("/admin/items/toggle")
async def toggle_item_activity(request: Request, item_id: int = Form(...), is_active: bool = Form(...)):
    cashier_id = get_cashier_id(request)
    # if cashier_id != "admin":
    #     return RedirectResponse("/", status_code=302)
    cashier = await conn.fetchrow("SELECT is_admin FROM cashiers WHERE id = $1", cashier_id)
    if not cashier or not cashier["is_admin"]:
        return RedirectResponse("/", status_code=302)


    async with request.app.state.db.acquire() as conn:
        await conn.execute("UPDATE items SET is_active = $1 WHERE id = $2", is_active, item_id)

    return RedirectResponse("/admin/items", status_code=302)


from datetime import (  # Import date and datetime for type hints and parsing
    date, datetime)


@router.get("/admin/orders", response_class=HTMLResponse)
async def admin_orders(request: Request):
    db = request.app.state.db

    order_for_date_str = request.query_params.get("order_for_date")
    address_filter = request.query_params.get("address")

    where_clauses: list = []
    # Use a list for parameters, as asyncpg expects them positionally for fetch
    query_args: list = []
    param_counter = 1 # To keep track of $1, $2, etc.

    if order_for_date_str:
        try:
            order_for_date = datetime.strptime(order_for_date_str, '%Y-%m-%d').date()
            where_clauses.append(f"o.order_for = ${param_counter}")
            query_args.append(order_for_date)
            param_counter += 1
        except ValueError:
            print(f"Warning: Invalid date format received: {order_for_date_str}")

    if address_filter:
        where_clauses.append(f"o.address ILIKE ${param_counter}")
        query_args.append(f"%{address_filter}%")
        param_counter += 1

    where_sql = ""
    if where_clauses:
        where_sql = " WHERE " + " AND ".join(where_clauses)

    async with db.acquire() as conn:
        rows = await conn.fetch(f"""
            SELECT
                o.id AS order_id,
                o.order_for,
                o.created,
                o.address,
                c.full_name AS cashier_name,
                oi.quantity,
                i.name AS item_name
            FROM orders o
            JOIN cashiers c ON o.cashier_id = c.id
            JOIN orders_items oi ON oi.order_id = o.id
            JOIN items i ON oi.item_id = i.id
            {where_sql}
            ORDER BY o.order_for DESC, o.created DESC
        """, *query_args)

    grouped_orders = {}
    for row in rows:
        date = row["order_for"]
        o_id = str(row["order_id"])
        if date not in grouped_orders:
            grouped_orders[date] = {}
        if o_id not in grouped_orders[date]:
            grouped_orders[date][o_id] = {
                "id": o_id,
                "created": row["created"],
                "address": row["address"],
                "cashier_name": row["cashier_name"],
                "items": [],
            }
        grouped_orders[date][o_id]["items"].append({
            "name": row["item_name"],
            "quantity": row["quantity"],
        })

    return templates.TemplateResponse("admin_orders.html", {
        "request": request,
        "grouped_orders": grouped_orders,
        "order_for": order_for_date_str,
        "address": address_filter
    })


@router.get("/admin/orders/export")
async def export_orders(address: str = None, request: Request = None):
    db = request.app.state.db
    async with db.acquire() as conn:
        query = """
            SELECT o.id, o.created, o.address, c.full_name AS cashier_name
            FROM orders o
            JOIN cashiers c ON o.cashier_id = c.id
        """
        params = []
        if address:
            query += " WHERE o.address ILIKE $1"
            params.append(f"%{address}%")
        query += " ORDER BY o.created DESC"

        orders = await conn.fetch(query, *params)
        logging.info(f"Orders fetched: {len(orders)}")
        if not orders:
            return {"detail": "No orders found"}

        wb = Workbook()
        wb.remove(wb.active)

        # Cache worksheets by address
        sheet_map = {}

        for order in orders:
            addr = order["address"] or "Unknown"
            sheet_name = addr[:31]  # Excel sheet name limit

            # Create or reuse sheet
            if sheet_name not in sheet_map:
                ws = wb.create_sheet(title=sheet_name)
                ws.append(["Order ID", "Created", "Address", "Cashier", "Item Name", "Quantity", "Price"])
                sheet_map[sheet_name] = ws
            else:
                ws = sheet_map[sheet_name]

            items = await conn.fetch("""
                SELECT oi.quantity, i.name, i.price
                FROM orders_items oi
                JOIN items i ON i.id = oi.item_id
                WHERE oi.order_id = $1
            """, order["id"])

            for item in items:
                ws.append([
                    str(order["id"]),
                    order["created"].strftime("%Y-%m-%d %H:%M:%S"),
                    order["address"],
                    order["cashier_name"],
                    item["name"],
                    item["quantity"],
                    float(item["price"])
                ])

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        headers = {
            "Content-Disposition": "attachment; filename=orders.xlsx"
        }

        return StreamingResponse(stream, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



@router.get("/admin/export/by_address")
async def export_by_address(order_for: date, request: Request):
    db = request.app.state.db
    async with db.acquire() as conn:
        query = """
            SELECT
                o.address,
                i.name AS item_name,
                oi.quantity
            FROM orders o
            JOIN orders_items oi ON o.id = oi.order_id
            JOIN items i ON oi.item_id = i.id
            WHERE o.order_for = $1
            ORDER BY o.address
        """
        rows = await conn.fetch(query, order_for)

    # Создаём Excel-файл
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Удаляем пустой первый лист

    # Группируем по адресу
    from collections import defaultdict
    grouped = defaultdict(list)
    for row in rows:
        grouped[row["address"]].append((row["item_name"], row["quantity"]))

    for address, items in grouped.items():
        ws = wb.create_sheet(title=address[:31])  # Excel max 31 chars
        ws.append(["Товар", "Количество"])
        for name, qty in items:
            ws.append([name, qty])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": f"attachment; filename=by_address_{order_for}.xlsx"
    })


@router.get("/admin/export/all_items")
async def export_all_items(order_for: date, request: Request):
    db = request.app.state.db
    async with db.acquire() as conn:
        query = """
            SELECT
                o.id AS order_id,
                i.name AS item_name,
                oi.quantity,
                o.address,
                o.order_for
            FROM orders o
            JOIN orders_items oi ON o.id = oi.order_id
            JOIN items i ON oi.item_id = i.id
            WHERE o.order_for = $1
            ORDER BY o.id, i.name;
        """
        rows = await conn.fetch(query, order_for)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Items by Order"
    
    # Set headers
    ws.append([
        "Order ID",
        "Name of Item",
        "Quantity",
        "Address",
        "Order Date"
    ])

    # Append data rows
    for row in rows:
        ws.append([
            str(row["order_id"]),  # <-- FIX: Convert UUID to string
            row["item_name"],
            row["quantity"],
            row["address"],
            row["order_for"]
        ])

    # Prepare response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=all_items_{order_for}.xlsx"
        }
    )


@router.post("/admin/orders/delete")
async def delete_order(request: Request, order_id: str = Form(...)):
    """
    Deletes an order and its associated items.
    Requires admin privileges.
    """
    cashier_id = get_cashier_id(request)
    db = request.app.state.db

    logging.info(f"Attempting to delete order with ID: '{order_id}' by cashier ID: {cashier_id}")

    # Defensive check: empty or missing order_id
    if not order_id or not order_id.strip():
        logging.error("No order_id provided in form submission.")
        return HTMLResponse("Missing order ID.", status_code=400)

    try:
        # Convert order_id to UUID
        order_uuid = uuid.UUID(order_id.strip())
        logging.info(f"Successfully parsed order_id to UUID: {order_uuid}")
    except ValueError:
        logging.error(f"Invalid UUID format for order_id: {order_id}")
        return HTMLResponse(f"Invalid order ID format: {order_id}", status_code=400)

    async with db.acquire() as conn:
        # Check if the user is an admin
        cashier = await conn.fetchrow("SELECT is_admin FROM cashiers WHERE id = $1", cashier_id)
        if not cashier or not cashier["is_admin"]:
            logging.warning(f"Unauthorized delete attempt by cashier ID: {cashier_id}")
            return RedirectResponse("/", status_code=302)

        try:
            # Start transaction
            async with conn.transaction():
                # Delete items
                items_deleted = await conn.execute("DELETE FROM orders_items WHERE order_id = $1", order_uuid)
                logging.info(f"Deleted from orders_items: {items_deleted}")

                # Delete order
                order_deleted = await conn.execute("DELETE FROM orders WHERE id = $1", order_uuid)
                logging.info(f"Deleted from orders: {order_deleted}")

            if order_deleted and "DELETE" in order_deleted:
                logging.info(f"Successfully deleted order {order_uuid}")
                return RedirectResponse("/admin/orders", status_code=302)
            else:
                logging.warning(f"No order deleted with ID {order_uuid}")
                return HTMLResponse("Order not found or already deleted.", status_code=404)

        except Exception as e:
            logging.error(f"Error deleting order {order_id}: {e}", exc_info=True)
            return HTMLResponse("Internal server error while deleting the order.", status_code=500)