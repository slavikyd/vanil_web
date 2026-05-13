import logging
from datetime import date, datetime
from io import BytesIO
from typing import Any

from django.contrib import admin
from django.db.models import Prefetch
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.template.response import TemplateResponse
from django.urls import path
from django.utils import timezone
from django.utils.dateparse import parse_date
from fastapi import status
from openpyxl import Workbook

from .models import *

logger = logging.getLogger(__name__)

PRIORITY_GROUP_NAME = "0-й рейс"
EXPORT_TABLE_GAP = 2

# TODO: consider move to plain SQL 

def _orders_payload(*, max_days: int | None, offset_days: int) -> dict[str, Any]:
    """
    Builds the full data payload for both live and archive views.

    max_days=5, offset_days=0  → newest 5 days  (live page)
    max_days=None, offset_days=5 → everything older (archive)
    """
    all_items = list(
        Items.objects
        .order_by("tbl", "pos", "name")
        .values("name", "tbl", "pos")
    )
    all_items = [i for i in all_items if i["name"]]

    all_shops = list(Shops.objects.order_by('address').values('id', 'address'))
    all_shops = [{'id': str(s['id']), 'address': s['address']} for s in all_shops]

    order_items_prefetch = (
        OrdersItems.objects
        .select_related("item")
        .only("order_id", "item_id", "quantity", 'order_type', "item__name")
    )
    orders_qs = (
        Orders.objects
        .select_related("cashier", "shop")
        .order_by("-order_for", "-created")
        .prefetch_related(Prefetch("ordersitems_set", queryset=order_items_prefetch))
    )

    grouped: dict[date, list] = {}
    for o in orders_qs:
        grouped.setdefault(o.order_for, []).append(o)

    today = timezone.localdate()
    days_sorted = sorted(grouped.keys(), reverse=True)[offset_days:]
    if max_days is not None:
        days_sorted = days_sorted[:max_days]

    days_payload = []
    for d in days_sorted:
        day_orders = sorted(
            grouped.get(d, []),
            key=lambda x: (x.created or datetime.min.replace(tzinfo=timezone.UTC)),
            reverse=True,
        )

        totals_shop: dict[str, int] = {i['name']: 0 for i in all_items}
        totals_special: dict[str, int] = {i['name']: 0 for i in all_items}
        totals_priority: dict[str, int] = {i['name']: 0 for i in all_items}
        shops_map: dict[str, list] = {}

        for o in day_orders:
            shop_id_key = str(getattr(o, 'shop_id', None) or o.address or 'unknown')
            shops_map.setdefault(shop_id_key, []).append(o)

            is_priority = (
                getattr(getattr(o, "shop", None), "shop_group", None) is not None
                and getattr(getattr(o.shop, "shop_group", None), "name", None) == PRIORITY_GROUP_NAME
            )

            for oi in o.ordersitems_set.all():
                name = getattr(oi.item, "name", None)
                if not name:
                    continue
                qty = int(oi.quantity or 0)
                if is_priority:
                    totals_priority[name] = totals_priority.get(name, 0) + qty
                elif getattr(oi, "order_type", "Обычный") == "Спец. заказ":
                    totals_special[name] = totals_special.get(name, 0) + qty
                else:
                    totals_shop[name] = totals_shop.get(name, 0) + qty



        totals_list = [
            {"name": i["name"], "quantity_shop": totals_shop.get(i["name"], 0), "quantity_special": totals_special.get(i['name'], 0), "quantity_priority": totals_priority.get(i["name"], 0), "tbl": i["tbl"], "pos": i["pos"]} # TODO: beautify
            for i in all_items
        ]

        shop_lookup = {str(s['id']): s['address'] for s in all_shops}
        shops_payload = []
        for shop_key in sorted(shops_map.keys()):
            shop_orders = sorted(
                shops_map[shop_key],
                key=lambda x: (x.created or datetime.min.replace(tzinfo=timezone.UTC)),
                reverse=True,
            )
            orders_payload = []
            for o in shop_orders:
                items_payload = [
                    {"name": oi.item.name, "quantity": int(oi.quantity or 0), "order_type": getattr(oi, "order_type", "Обычный")}
                    for oi in o.ordersitems_set.all()
                    if getattr(oi.item, "name", None)
                ]
                orders_payload.append({
                    "id": str(o.id),
                    "created": o.created.isoformat() if o.created else None,
                    "created_hhmm": o.created.strftime("%H:%M:%S") if o.created else "",
                    "address": o.address,
                    "cashier_name": getattr(o.cashier, "full_name", "") if o.cashier_id else "",
                    "shop_id": getattr(o, "shop_id", None),
                    "items": items_payload,
                    'disabled': o.disabled,
                    'completed': o.completed,
                })
            
            shops_payload.append({"shop_id": shop_key, "shop": shop_lookup.get(shop_key, shop_key), "orders": orders_payload})

        days_payload.append({
            "date": d.isoformat(),
            "is_today": d == today,
            "total_orders": len(day_orders),
            "totals": totals_list,
            "shops": shops_payload,
        })

    return {
        "generated_at": timezone.now().isoformat(),
        'all_shops': all_shops,
        "days": days_payload,
    }



@admin.register(Orders)
class OrdersAdmin(admin.ModelAdmin):
    change_list_template = "admin/orders/live.html"

    def print_totals_view(self, request: HttpRequest) -> HttpResponse:
        order_for = parse_date(request.GET.get("order_for") or "")
        if not order_for:
            return HttpResponse("Missing ?order_for=YYYY-MM-DD", status=400)

        all_items = list(
            Items.objects.order_by("tbl", 'pos', 'name').values('name', 'tbl', 'pos')
        )
        all_items = [i for i in all_items if i['name']]

        totals_plain: dict[str, int] = {i['name']: 0 for i in all_items}
        totals_special: dict[str, int] = {i['name']: 0 for i in all_items}
        totals_priority: dict[str, int] = {i['name']: 0 for i in all_items}

        for oi in (OrdersItems.objects
                .select_related("item", 'order__shop__shop_group')
                .filter(order__order_for=order_for)):
            name = getattr(oi.item, "name", None)
            if not name:
                continue
            qty = int(oi.quantity or 0)
            order_type = getattr(oi, 'order_type', 'Обычный')
            is_priority = (
                getattr(getattr(getattr(oi, 'order', None), 'shop', None), 'shop_group', None) is not None
                and getattr(oi.order.shop.shop_group, 'name', None) == PRIORITY_GROUP_NAME
            )
            if is_priority:
                totals_priority[name] = totals_priority.get(name, 0) + qty
            elif order_type == 'Спец. заказ':
                totals_special[name] = totals_special.get(name, 0) + qty
            else:
                totals_plain[name] = totals_plain.get(name, 0) + qty

        tbl_groups: dict[int, list] = {}
        for item in all_items:
            tbl = item['tbl'] if item['tbl'] is not None else 3
            tbl_groups.setdefault(tbl, []).append(item)

        tables = []
        for tbl_index in range(4):
            items = tbl_groups.get(tbl_index, [])
            tables.append([
                {
                    'name': i['name'],
                    'plain': totals_plain.get(i['name'], 0),
                    'special': totals_special.get(i['name'], 0),
                    'priority': totals_priority.get(i['name'], 0),
                }
                for i in items
            ])

        ctx = {
            'order_for': order_for,
            'tables': tables,
        }
        return TemplateResponse(request, "admin/orders/print_totals.html", ctx)

    def print_shop_view(self, request: HttpRequest) -> HttpResponse:
        order_for = parse_date(request.GET.get("order_for") or "")
        shop_id = request.GET.get("shop_id") or ""
        if not order_for or not shop_id:
            return HttpResponse("Missing ?order_for=YYYY-MM-DD&shop_id=...", status=400)

        try:
            shop = Shops.objects.get(id=shop_id)
        except Shops.DoesNotExist:
            return HttpResponse("Shop not found", status=404)

        all_items = list(
            Items.objects.order_by("tbl", 'pos', 'name').values('name', 'tbl', 'pos')
        )
        all_items = [i for i in all_items if i['name']]

        totals_special: dict[str, int] = {i['name']: 0 for i in all_items}
        totals_plain: dict[str, int] = {i['name']: 0 for i in all_items}
        item_comments: dict[str, str] = {}
        order_comments: list[str] = []

        orders = (
            Orders.objects
            .filter(shop_id=shop_id, order_for=order_for)
            .prefetch_related(
                Prefetch(
                    "ordersitems_set",
                    queryset=OrdersItems.objects.select_related("item").order_by("item__tbl", "item__pos", "item__name")
                )
            )
            .order_by("created")
        )

        for order in orders:
            if order.comment:
                order_comments.append(order.comment)
            for oi in order.ordersitems_set.all():
                name = getattr(oi.item, "name", None)
                if not name:
                    continue
                qty = int(oi.quantity or 0)
                order_type = getattr(oi, "order_type", "Обычный")
                if order_type == "Спец. заказ":
                    totals_special[name] = totals_special.get(name, 0) + qty
                else:
                    totals_plain[name] = totals_plain.get(name, 0) + qty
                if oi.comment:
                    item_comments[name] = oi.comment

        tbl_groups: dict[int, list] = {}
        for item in all_items:
            tbl = item['tbl'] if item['tbl'] is not None else 3
            tbl_groups.setdefault(tbl, []).append(item)

        tables = []
        for tbl_index in range(4):
            items = tbl_groups.get(tbl_index, [])
            tables.append([
                {
                    'name': i['name'],
                    'plain': totals_plain.get(i['name'], 0),
                    'special': totals_special.get(i['name'], 0),
                    'comment': item_comments.get(i['name'], ''),
                }
                for i in items
            ])

        ctx = {
            'order_for': order_for,
            'shop': shop,
            'tables': tables,
            'order_comments': order_comments,
        }
        return TemplateResponse(request, "admin/orders/print_shop.html", ctx)

    def get_urls(self):

        urls = super().get_urls()
        custom = [
            path("live-data/",
                 self.admin_site.admin_view(self.live_data_view),
                 name="orders_live_data"),

            path("archive/",
                 self.admin_site.admin_view(self.archive_view),
                 name="orders_archive"),

            path("archive/data/",
                 self.admin_site.admin_view(self.archive_data_view),
                 name="orders_archive_data"),

            path("export/totals/",
                 self.admin_site.admin_view(self.export_totals_view),
                 name="orders_export_totals"),

            path("export/shop/",
                self.admin_site.admin_view(self.export_shop_view),
                name="orders_export_shop"),

            path("order/<str:order_id>/toggle-disabled/",
                self.admin_site.admin_view(self.toggle_disabled_view),
                name="orders_toggle_disabled"),

            path("order/<str:order_id>/toggle-completed/",
                self.admin_site.admin_view(self.toggle_completed_view),
                name="orders_toggle_completed"),

            path("order/<str:order_id>/delete/",
                self.admin_site.admin_view(self.delete_order_view),
                name="orders_delete_order"),

            path("print/totals/",
                self.admin_site.admin_view(self.print_totals_view),
                name="orders_print_totals"),

            path("print/shop/",
                self.admin_site.admin_view(self.print_shop_view),
                name="orders_print_shop"),
        ]

        return custom + urls

    def changelist_view(self, request, extra_context=None):

        ctx = {
            **self.admin_site.each_context(request),
            "title": "Заявки",
            "data_url": "live-data/",
            "archive_url": "archive/",
            "export_url": "export/totals/",
            "show_archive_link": True,
            "archive_heading": "Архив",
            "opts": self.model._meta,
            **(extra_context or {}),
        }
        return TemplateResponse(request, "admin/orders/live.html", ctx)

    def live_data_view(self, request: HttpRequest) -> JsonResponse:
        """Returns JSON for the live page (newest 5 days)."""
        payload = _orders_payload(max_days=5, offset_days=0)
        return JsonResponse(payload)
    
    def delete_order_view(self, request: HttpRequest, order_id: str) -> JsonResponse:
        if request.method != 'POST':
            return JsonResponse({'error': 'POST required'}, status=405)
        try:
            order = Orders.objects.get(id=order_id)
            OrdersItems.objects.filter(order=order).delete()
            order.delete()
            return JsonResponse({'deleted': True})
        except Orders.DoesNotExist:
            return JsonResponse({'error': 'not found'}, status=404)

    def archive_view(self, request: HttpRequest) -> HttpResponse:
        """Renders the archive HTML page (orders older than 5 days)."""
        ctx = {
            **self.admin_site.each_context(request),
            "title": "Архив заявок",
            "data_url": "data/",
            "live_url": "../",
            "show_live_link": True,
            "opts": self.model._meta,
        }
        return TemplateResponse(request, "admin/orders/archive.html", ctx)

    def toggle_disabled_view(self, request: HttpRequest, order_id: str) -> JsonResponse:
        if request.method != 'POST':
            return JsonResponse({'error': 'POST required'}, status=status.HTTP_405_METHOD_NOT_ALLOWED )
        try:
            order = Orders.objects.get(id=order_id)
            order.disabled = not order.disabled
            order.save(update_fields=['disabled'])
            return JsonResponse({'disabled': order.disabled})
        except Orders.DoesNotExist:
            return JsonResponse({'error': 'not found'}, status=status.HTTP_404_NOT_FOUND)

    def toggle_completed_view(self, request: HttpRequest, order_id: str) -> JsonResponse:
        if request.method != 'POST':
            return JsonResponse({'error': 'POST required'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)
        try:
            order = Orders.objects.get(id=order_id)
            order.completed = not order.completed
            order.save(update_fields=['completed'])
            return JsonResponse({'completed': order.completed})
        except Orders.DoesNotExist:
            return JsonResponse({'error': 'not found'}, status=status.HTTP_404_NOT_FOUND)

    def archive_data_view(self, request: HttpRequest) -> JsonResponse:
        """Returns JSON for the archive page (everything older than 5 days)."""
        payload = _orders_payload(max_days=None, offset_days=5)
        return JsonResponse(payload)

    def export_totals_view(self, request: HttpRequest) -> HttpResponse:
        order_for = parse_date(request.GET.get("order_for") or "")
        if not order_for:
            return HttpResponse("Missing ?order_for=YYYY-MM-DD", status=status.HTTP_400_BAD_REQUEST)

        all_items = list(
            Items.objects.order_by("tbl", 'pos', 'name').values('name', 'tbl', 'pos')
        )
        all_items = [i for i in all_items if i['name']]

        totals_plain: dict[str, int] = {i['name']: 0 for i in all_items}
        totals_special: dict[str, int] = {i['name']: 0 for i in all_items}
        totals_priority: dict[str, int] = {i['name']: 0 for i in all_items}

        for oi in (OrdersItems.objects
                .select_related("item", 'order__shop__shop_group')
                .filter(order__order_for=order_for)):
            name = getattr(oi.item, "name", None)
            if not name:
                continue
            qty = int(oi.quantity or 0)
            order_type = getattr(oi, 'order_type', 'Обычный')
            is_priority = (
                getattr(getattr(getattr(oi, 'order', None), 'shop', None), 'shop_group', None) is not None
                and getattr(oi.order.shop.shop_group, 'name', None) == PRIORITY_GROUP_NAME
            )
            if is_priority:
                totals_priority[name] = totals_priority.get(name, 0) + qty
            elif order_type == 'Спец. заказ':
                totals_special[name] = totals_special.get(name, 0) + qty
            else:
                totals_plain[name] = totals_plain.get(name, 0) + qty

        tbl_groups: dict[int, list] = {}
        for item in all_items:
            tbl = item['tbl'] if item['tbl'] is not None else 3
            tbl_groups.setdefault(tbl, []).append(item)

        DATA_COLS = 4
        right_col_start = DATA_COLS + EXPORT_TABLE_GAP + 1

        curr_date = datetime.now()
        SHEET_NAME = f"Total_{curr_date.year}-{curr_date.month}-{curr_date.day}_{curr_date.hour}-{curr_date.minute}"
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

        def write_table(tbl_index: int, start_row: int, start_col: int) -> int:
            items = tbl_groups.get(tbl_index, [])
            if not items:
                return start_row
            ws.cell(row=start_row, column=start_col, value='Позиция')
            ws.cell(row=start_row, column=start_col + 1, value='Заказ с магазина')
            ws.cell(row=start_row, column=start_col + 2, value='Заказ')
            ws.cell(row=start_row, column=start_col + 3, value='Юбик')
            start_row += 1
            for item in items:
                name = item['name']
                ws.cell(row=start_row, column=start_col, value=name)
                ws.cell(row=start_row, column=start_col + 1, value=totals_plain.get(name, 0))
                ws.cell(row=start_row, column=start_col + 2, value=totals_special.get(name, 0))
                ws.cell(row=start_row, column=start_col + 3, value=totals_priority.get(name, 0))
                start_row += 1
            return start_row + 1 

        left_row = 1
        left_row = write_table(0, left_row, 1)
        left_row = write_table(1, left_row, 1)

        right_row = 1
        right_row = write_table(2, right_row, right_col_start)
        right_row = write_table(3, right_row, right_col_start)

        buf = BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        logger.info('Export of main table', extra={'order_for': str(order_for)})
        resp['Content-Disposition'] = f'attachment; filename="{SHEET_NAME}.xlsx"'
        return resp
    
    def export_shop_view(self, request: HttpRequest) -> HttpResponse:
        order_for = parse_date(request.GET.get("order_for") or "")
        shop_id = request.GET.get("shop_id") or ""
        if not order_for or not shop_id:
            return HttpResponse("Missing ?order_for=YYYY-MM-DD&shop_id=...", status=400)

        try:
            shop = Shops.objects.get(id=shop_id)
        except Shops.DoesNotExist:
            return HttpResponse("Shop not found", status=404)

        all_items = list(
            Items.objects.order_by("tbl", 'pos', 'name').values('name', 'tbl', 'pos')
        )
        all_items = [i for i in all_items if i['name']]

        totals_special: dict[str, int] = {i['name']: 0 for i in all_items}
        totals_plain: dict[str, int] = {i['name']: 0 for i in all_items}
        item_comments: dict[str, str] = {}
        order_comments: list[str] = []

        orders = (
            Orders.objects
            .filter(shop_id=shop_id, order_for=order_for)
            .prefetch_related(
                Prefetch(
                    "ordersitems_set",
                    queryset=OrdersItems.objects.select_related("item").order_by("item__tbl", "item__pos", "item__name")
                )
            )
            .order_by("created")
        )

        for order in orders:
            if order.comment:
                order_comments.append(order.comment)
            for oi in order.ordersitems_set.all():
                name = getattr(oi.item, "name", None)
                if not name:
                    continue
                qty = int(oi.quantity or 0)
                order_type = getattr(oi, "order_type", "Обычный")
                if order_type == "Спец. заказ":
                    totals_special[name] = totals_special.get(name, 0) + qty
                else:
                    totals_plain[name] = totals_plain.get(name, 0) + qty
                if oi.comment:
                    item_comments[name] = oi.comment

        # Group items by tbl
        tbl_groups: dict[int, list] = {}
        for item in all_items:
            tbl = item['tbl'] if item['tbl'] is not None else 3
            tbl_groups.setdefault(tbl, []).append(item)

        DATA_COLS = 4  # Позиция + Спец. заказ + Обычный + Комментарий
        right_col_start = DATA_COLS + EXPORT_TABLE_GAP + 1

        wb = Workbook()
        ws = wb.active
        ws.title = (shop.address or shop_id)[:31]  # Excel sheet name limit

        def write_table(tbl_index: int, start_row: int, start_col: int) -> int:
            items = tbl_groups.get(tbl_index, [])
            if not items:
                return start_row
            ws.cell(row=start_row, column=start_col, value='Позиция')
            ws.cell(row=start_row, column=start_col + 1, value='Спец. заказ')
            ws.cell(row=start_row, column=start_col + 2, value='Обычный')
            ws.cell(row=start_row, column=start_col + 3, value='Комментарий')
            start_row += 1
            for item in items:
                name = item['name']
                ws.cell(row=start_row, column=start_col, value=name)
                ws.cell(row=start_row, column=start_col + 1, value=totals_special.get(name, 0))
                ws.cell(row=start_row, column=start_col + 2, value=totals_plain.get(name, 0))
                ws.cell(row=start_row, column=start_col + 3, value=item_comments.get(name, ''))
                start_row += 1
            return start_row + 1

        # Left side: tables 0 and 1
        left_row = 1
        left_row = write_table(0, left_row, 1)
        left_row = write_table(1, left_row, 1)

        # Right side: tables 2 and 3
        right_row = 1
        right_row = write_table(2, right_row, right_col_start)
        right_row = write_table(3, right_row, right_col_start)

        # Order comments at the bottom
        bottom_row = max(left_row, right_row) + 1
        for comment in order_comments:
            ws.cell(row=bottom_row, column=1, value='Комментарий к заказу:')
            ws.cell(row=bottom_row, column=2, value=comment)
            bottom_row += 1

        buf = BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        logger.info('Export of shop table', extra={'shop_id': shop_id, 'order_for': str(order_for)})
        resp["Content-Disposition"] = f'attachment; filename="shop_{shop_id}_{order_for}.xlsx"'
        return resp




@admin.register(Cashiers)
class CashiersAdmin(admin.ModelAdmin):
    list_display = [f.name for f in Cashiers._meta.fields]

@admin.register(Categories)
class CategoriesAdmin(admin.ModelAdmin):
    list_display = [f.name for f in Categories._meta.fields]

@admin.register(Items)
class ItemsAdmin(admin.ModelAdmin):
    list_display = ["name", "active", "category", 'tbl', 'pos']
    list_editable = ["active", "category", 'tbl', 'pos']

@admin.register(Shops)
class ShopsAdmin(admin.ModelAdmin):
    list_display = ['id', 'address', 'shop_group']
    list_editable = ['shop_group']
    list_display_links = ['id']

@admin.register(ShopsGroups)
class ShopsGroupsAdmin(admin.ModelAdmin):
    list_display = ['id', 'name']